#!/usr/bin/python

import argparse
import csv
import sys

from openpyxl import load_workbook


IGNORE_CASE = True
LOWER_RESULTS = True

def error(msg, and_exit=True):
    print "ERROR: %s" % msg
    if and_exit:
        sys.exit(-1)


class Analyzer(object):

    def __init__(self, headers=["JOB", "VALUE", "PARENT"]):
        self._headers = headers
        self._data = []
        self._analyzed = False

    def analyze_csv(self, filename):
        data = []
        with open(filename) as f:
            csv_reader = csv.DictReader(f, self._headers, delimiter=";")
            csv_reader.next()  # Skip first "header" line
            data = [row for row in csv_reader] 
            self._analyzed = True
        self._data = data

    def analyze_excel(self, filename):
        data = []
        wb = load_workbook(filename, read_only=True)
        ws = wb.get_sheet_by_name('Sheet1')
        for row in ws.iter_rows(row_offset=1):
            entry = {}
            for index, header in enumerate(self._headers):
                entry[header] = row[index].value
            data.append(entry)
        self._analyzed = True
        self._data = data                

    def find_job(self, name):
        if not self._analyzed:
            error("before you can find a job you have load/analyze a source file (csv or excel)")
        matches = []
        if IGNORE_CASE:
            matches = [hit for hit in self._data if hit['JOB'].lower() == name.lower()]
        else:
            matches = [hit for hit in self._data if hit['JOB'] == name]
        if not matches:
            return None
        if len(matches) > 1:
            error("Too many matches for job '%s'" % name)
        return matches[0]

    def find_job_chain(self, name):
        result = []
        found_job_names = []
        search_job = name
        while True:
            job = self.find_job(search_job)
            if job:
                result.append(job)
                found_job_names.append(job['JOB'])
                parent = job.get('PARENT', None)
                if not parent or parent in found_job_names:
                    break
                search_job = parent
            else:
                break
        return result


def main():
    print ""
    parser = argparse.ArgumentParser(description="Job analyzer")
    parser.add_argument('-p', dest='job', help='job name')
    parser.add_argument('-o', dest='logfile', help='additional logfile to log result into')

    args = parser.parse_args()
    job_to_analyze = args.job
    if not job_to_analyze:
        parser.print_usage()
        error("missing job defined by parameter '-p'")

    job_analyzer = Analyzer()
    #job_analyzer.analyze_csv("INPUT.csv")
    job_analyzer.analyze_excel("INPUT.xlsx")
    
    result = []
    chain = job_analyzer.find_job_chain(job_to_analyze)
    for i, elem in enumerate(chain):
        result.append(elem['JOB'].lower() if LOWER_RESULTS else elem['JOB'])
        if i >= len(chain)-1 and elem['PARENT']:
            result.append(elem['PARENT'].lower() if LOWER_RESULTS else elem['PARENT'])

    out = "    ".join(result)
    print out

    if args.logfile:
        try:
            with open(args.logfile, 'w') as log:
                log.write("%s\n" % out)
        except Exception as err:
            error("Could not write logfile '%s': %s" % (args.logfile, err))


if __name__ == '__main__':
    main()
